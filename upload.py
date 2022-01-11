# Below importing module related to Google sheet
from pprint import pprint
from googleapiclient import discovery
from oauth2client.service_account import ServiceAccountCredentials

# Below importing module related to excel file
import openpyxl

# Below importing module related to airflow
import airflow
from airflow import DAG
from airflow.operators.dummy import DummyOperator
from airflow.operators.python import PythonOperator
from airflow.operators.bash import BashOperator

from datetime import datetime
from datetime import timedelta

# Process airflow
default_args = {
    'owner': 'default_user',
    'start_date': airflow.utils.dates.days_ago(1),
    'depends_on_past': True,
    'email': ['dodinovembri32@gmail.com'],
    'email_on_failure': True,
    'email_on_retry': True,
    'retries': 5,
    'retry_delay': timedelta(minutes=30),
}

dag = DAG(
    'sheets_pipeline',
    default_args=default_args,
    schedule_interval=timedelta(days=1),
)


# Start 
scope = ["https://spreadsheets.google.com/feeds", 'https://www.googleapis.com/auth/spreadsheets',
         "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]
credentials = ServiceAccountCredentials.from_json_keyfile_name('key/client_secret.json', scope)
service = discovery.build('sheets', 'v4', credentials=credentials)

# The ID of the spreadsheet to update.
spreadsheet_id = '1Y6-i90LU33uJcxRxoVPXBFDPwq2UQwh8tahTWMTLPrE'

# The A1 notation of a range to search for a logical table of data.
# Values will be appended after the last row of the table.
range_ = 'A1:H1'

# How the input data should be interpreted.
value_input_option = 'RAW'

# How the input data should be inserted.
insert_data_option = 'OVERWRITE'

# start function -----------------------------------------------------------------------------------------------------------------

def my_func():
    # load excel with its path
    wrkbk = openpyxl.load_workbook("Data Penjualan CV Fortuna.xlsx", data_only = True)
    sh = wrkbk.active

    # -----------------------------------------------------------------------------

    # Extract Process - iterate through excel and display data
    values = []
    for i in range(1, sh.max_row+1):
        values_each = []
        for j in range(1, sh.max_column+1):
            cell_obj = sh.cell(row=i, column=j)
            if cell_obj.value == None:
                continue
            values_each.append(cell_obj.value)

        values.append(values_each)

    # -----------------------------------------------------------------------------

    # Transform Process 

    # 1. Find not empty row
    values_fix = [x for x in values if x != []]


    value_range_body = {
        "majorDimension": "ROWS",
        "values": values_fix
    }

    # -----------------------------------------------------------------------------

    # Delete existing
    range_delete = '!A1:Z'
    clear_values_request_body = {}
    request = service.spreadsheets().values().clear(spreadsheetId=spreadsheet_id, range=range_delete, body=clear_values_request_body)
    response = request.execute()

    # Load Process - load to google sheet
    request = service.spreadsheets().values().append(spreadsheetId=spreadsheet_id, range=range_, valueInputOption=value_input_option, insertDataOption=insert_data_option, body=value_range_body)

    # -----------------------------------------------------------------------------

    # Process response
    response = request.execute()
    pprint(response)

# end function -------------------------------------------------------------------------------------------------------------------------

# run airflow

bashtask        = BashOperator(
                task_id='print_date',
                bash_command='date',
                dag=dag)

dummy_task      = DummyOperator(
                task_id='dummy_task',
                retries=3,
                dag=dag)

python_task     = PythonOperator(
                task_id='python_task',
                python_callable=my_func,
                dag=dag)

bashtask.set_downstream(python_task)

dummy_task.set_upstream(bashtask)